print("This code was made by Verygafanhot ( https://github.com/Verygafanhot )\nWaiting for next scheduled run")

import smtplib
from openpyxl import load_workbook
import datetime
import asyncio
import schedule
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Load the excel file
data = load_workbook('birthday.xlsx')
ws = data.active

async def async_while_true(today_day_month):
    Number = 0
    # check every line until "END"
    while True:
        Number += 1
        DiaExel = ws[f'C{Number}']
        if DiaExel.value is None:
            continue

        TodayExcel = str(DiaExel.value)
        try:
            excel_date = datetime.datetime.strptime(TodayExcel, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            if TodayExcel == "END":
                print("Ended, waiting until tomorrow!!\n-------\n\n")
                break
            continue

        excel_day_month = excel_date.strftime('%m-%d')

        if excel_day_month == today_day_month:
            print("Birthday detected... Getting info...")
            NomeExel = ws[f'B{Number}']
            MailExel = ws[f'D{Number}']
            NomeExelStr = str(NomeExel.value)
            MailExelStr = str(MailExel.value)
            print(f'Name= {NomeExelStr}. Mail={MailExelStr}\nSending email...')

            # Send email
            subject = f'Happy Birthday {NomeExelStr}'
            # HTML body with image included
            #Change the image url and text as you desire
            body = """
            <html>
            <body>
            <p>Happy Birthday!</p>
            <img src="https://t3.ftcdn.net/jpg/04/42/62/12/360_F_442621279_PYhie13pVGcSSYTAm1eqlC3e7Lcy0oNV.jpg" alt="Happy Birthday Image">
            </body>
            </html>
            """
            sender = "birthdays@example.com" # The sender email
            recipients = [MailExelStr]
            password = "0000 0000 0000 0000"  # Get it here https://myaccount.google.com/apppasswords (2fa required)

            async def send_email(subject, body, sender, recipients, password):
                msg = MIMEMultipart('alternative')
                msg['Subject'] = subject
                msg['From'] = sender
                msg['To'] = ', '.join(recipients)

                part1 = MIMEText(body, 'html')
                msg.attach(part1)

                with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp_server:
                    smtp_server.login(sender, password)
                    smtp_server.sendmail(sender, recipients, msg.as_string())

            await send_email(subject, body, sender, recipients, password)
            print("Mail sent! Looking for more...\n\n")


async def main():
    today = datetime.date.today()
    today_day_month = today.strftime('%m-%d')
    await async_while_true(today_day_month)


def job():
    asyncio.run(main())


# Agendar o trabalho para 8AM todos os dias
schedule.every().day.at("08:00").do(job)

while True:
    schedule.run_pending()
    time.sleep(1)
